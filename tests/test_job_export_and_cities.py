import csv
import io
import json
import tempfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from jobagent.cities import CityRefreshError, load_city_snapshot, refresh_city_cache
from jobagent.db import get_db, insert_job, update_job_status
from jobagent.job_export import InvalidJobSelectionError, build_csv, build_xlsx, export_jobs


def _job(job_id: str, *, city: str = "北京", company: str = "Example") -> dict:
	return {
		"id": job_id,
		"title": "Engineer",
		"company": company,
		"salary": "10-20K",
		"city": city,
		"experience": "1-3 years",
		"jd": "Build product features",
		"hr_name": "HR",
		"hr_title": "Recruiter",
		"hr_active": "active",
		"company_size": "100-499",
		"company_industry": "Software",
		"url": f"https://example.com/jobs/{job_id}",
	}


def _csv_ids(content: bytes) -> set[str]:
	rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
	id_column = rows[0].index("岗位 ID")
	return {row[id_column] for row in rows[1:]}


def test_bundled_city_snapshot_is_complete_and_offline(tmp_path):
	with patch("jobagent.cities.httpx.get") as http_get:
		snapshot = load_city_snapshot(cache_path=tmp_path / "cities.cache.json")

	assert snapshot["source"] == "bundled"
	assert len(snapshot["cities"]) >= 300
	codes = {city["name"]: city["code"] for city in snapshot["cities"]}
	assert codes.items() >= {
		"北京": "101010100",
		"上海": "101020100",
		"广州": "101280100",
		"深圳": "101280600",
	}.items()
	http_get.assert_not_called()


def test_invalid_city_cache_falls_back_without_network(tmp_path):
	cache_path = tmp_path / "cities.cache.json"
	cache_path.write_text("<html>not-json</html>", encoding="utf-8")
	with patch("jobagent.cities.httpx.get") as http_get:
		snapshot = load_city_snapshot(cache_path=cache_path)

	assert snapshot["source"] == "bundled"
	http_get.assert_not_called()


def test_html_city_refresh_does_not_overwrite_cache(tmp_path):
	cache_path = tmp_path / "cities.cache.json"
	original = load_city_snapshot(cache_path=cache_path)
	cache_path.write_text(json.dumps(original, ensure_ascii=False), encoding="utf-8")
	response = SimpleNamespace(
		status_code=200,
		headers={"content-type": "text/html"},
		text="<html>blocked</html>",
	)

	with pytest.raises(CityRefreshError):
		refresh_city_cache(cache_path, fetcher=lambda *args, **kwargs: response)

	assert json.loads(cache_path.read_text(encoding="utf-8"))["schema"] == "jobagent.cities.v1"


def test_export_scopes_use_exact_database_sets(tmp_path):
	db = get_db(tmp_path / "export.db")
	try:
		for index in range(8):
			job = _job(f"range-{index}", city="北京" if index < 4 else "上海")
			job["title"] = "Python 后端" if index in {0, 1, 2} else "Java 后端"
			insert_job(db, job)
			if index in {0, 1, 2}:
				update_job_status(db, job["id"], "ready")

		all_content, _, _ = export_jobs(db, format="csv", scope="all")
		filtered_content, _, _ = export_jobs(
			db,
			format="csv",
			scope="filtered",
			filters={"q": "Python", "city": "北京", "status": "ready"},
		)
		selected_content, _, _ = export_jobs(
			db,
			format="csv",
			scope="selected",
			job_ids=["range-1", "range-7"],
		)
	finally:
		db.close()

	assert _csv_ids(all_content) == {f"range-{index}" for index in range(8)}
	assert _csv_ids(filtered_content) == {"range-0", "range-1", "range-2"}
	assert _csv_ids(selected_content) == {"range-1", "range-7"}


def test_export_supports_51job_source_filter_and_label(tmp_path):
	db = get_db(tmp_path / "job51-export.db")
	try:
		job = _job("51job:sh-1", city="上海")
		job.update({
			"source_platform": "51job",
			"source_job_id": "sh-1",
			"source_city_code": "020000",
		})
		insert_job(db, job)
		content, _, _ = export_jobs(
			db,
			format="csv",
			scope="filtered",
			filters={"source_platform": "51job"},
		)
	finally:
		db.close()

	text = content.decode("utf-8-sig")
	assert "前程无忧" in text
	assert "020000" in text


def test_selected_export_rejects_missing_ids(tmp_path):
	db = get_db(tmp_path / "missing.db")
	try:
		insert_job(db, _job("existing"))
		with pytest.raises(InvalidJobSelectionError) as error:
			export_jobs(db, format="csv", scope="selected", job_ids=["existing", "missing"])
	finally:
		db.close()

	assert error.value.invalid_ids == ["missing"]


def test_csv_has_bom_formula_protection_and_city_code():
	job = _job("csv-1", company="=2+2")
	job["salary"] = "-1"
	content = build_csv([job])

	assert content.startswith(b"\xef\xbb\xbf")
	text = content.decode("utf-8-sig")
	assert "'=2+2" in text
	assert "'-1" in text
	assert "101010100" in text


def test_xlsx_contains_clickable_job_url():
	job = _job("xlsx-1", city="上海")
	workbook = load_workbook(io.BytesIO(build_xlsx([job])), read_only=False)
	try:
		sheet = workbook.active
		url_column = next(cell.column for cell in sheet[1] if cell.value == "岗位链接")
		url_cell = sheet.cell(2, url_column)
		assert url_cell.value == "https://example.com/jobs/xlsx-1"
		assert url_cell.hyperlink.target == "https://example.com/jobs/xlsx-1"
	finally:
		workbook.close()

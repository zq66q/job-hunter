"""In-memory CSV/XLSX export for database-backed job selections."""

from __future__ import annotations

import csv
from copy import copy
import io
import json
import sqlite3
from typing import Any, Iterable

from jobagent.cities import get_city_code
from jobagent.job_filters import parse_monthly_salary_k


MAX_EXPORT_JOB_IDS = 1000


EXPORT_COLUMNS = [
	("岗位 ID", "id"),
	("来源平台", "source_platform_label"),
	("来源岗位 ID", "source_job_id"),
	("来源关键词", "source_keyword"),
	("职位", "title"),
	("公司", "company"),
	("薪资", "salary"),
	("城市", "city"),
	("平台城市编码", "platform_city_code"),
	("工作经验", "experience"),
	("学历要求", "education"),
	("招聘类型", "recruitment_type_label"),
	("JD", "jd"),
	("HR 姓名", "hr_name"),
	("HR 职位", "hr_title"),
	("HR 活跃度", "hr_active"),
	("公司规模", "company_size"),
	("公司行业", "company_industry"),
	("岗位链接", "url"),
	("预筛分", "quick_score"),
	("AI 分数", "score"),
	("评分理由", "score_reason"),
	("过滤来源", "filter_source"),
	("过滤原因", "filter_reason"),
	("人工推进", "manual_override"),
	("当前状态", "status"),
	("招呼语", "greeting"),
	("招呼语状态", "greeting_status"),
	("简历路径", "resume_path"),
	("创建时间", "created_at"),
	("更新时间", "updated_at"),
	("最近评分失败原因", "score_failure_reason"),
	("最近招呼语失败原因", "greeting_failure_reason"),
]


class InvalidJobSelectionError(ValueError):
	"""Raised when a selected export contains missing or deleted IDs."""

	def __init__(self, invalid_ids: list[str]):
		self.invalid_ids = invalid_ids
		super().__init__("所选岗位中存在无效、已删除或已过期的岗位 ID")


def _normalize_export_ids(job_ids: Iterable[Any] | None) -> list[str]:
	if job_ids is None:
		return []
	if isinstance(job_ids, (str, bytes, dict)):
		raise ValueError("岗位 ID 必须是数组")
	try:
		values = list(job_ids)
	except TypeError as exc:
		raise ValueError("岗位 ID 必须是数组") from exc
	ids: list[str] = []
	for value in values:
		if not isinstance(value, str):
			raise ValueError("岗位 ID 必须是字符串")
		job_id = value.strip()
		if job_id and job_id not in ids:
			ids.append(job_id)
	if len(ids) > MAX_EXPORT_JOB_IDS:
		raise ValueError(f"一次最多导出 {MAX_EXPORT_JOB_IDS} 个岗位")
	return ids


def _filtered_rows(conn: sqlite3.Connection, filters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
	filters = filters or {}
	conditions: list[str] = ["deleted_at IS NULL"]
	params: list[Any] = []
	keyword = str(filters.get("q") or filters.get("query") or "").strip()
	if keyword:
		conditions.append("(title LIKE ? OR company LIKE ? OR jd LIKE ? OR score_reason LIKE ?)")
		params.extend([f"%{keyword}%"] * 4)
	city = str(filters.get("city") or "").strip()
	if city:
		conditions.append("city = ?")
		params.append(city)
	status = str(filters.get("status") or "").strip()
	if status:
		conditions.append("status = ?")
		params.append(status)
	source_platform = str(filters.get("source_platform") or "").strip()
	if source_platform:
		if source_platform not in {"boss", "zhilian", "51job"}:
			raise ValueError("source_platform 参数无效")
		conditions.append("COALESCE(source_platform, 'boss') = ?")
		params.append(source_platform)
	recruitment_type = str(filters.get("recruitment_type") or "").strip()
	if recruitment_type:
		if recruitment_type not in {"campus", "experienced", "unknown"}:
			raise ValueError("recruitment_type 参数无效")
		conditions.append("COALESCE(recruitment_type, 'unknown') = ?")
		params.append(recruitment_type)
	education = str(filters.get("education") or "").strip()
	if education:
		conditions.append("education LIKE ?")
		params.append(f"%{education}%")
	minimum_score = filters.get("min_score")
	if minimum_score not in (None, ""):
		try:
			minimum_score = float(minimum_score)
		except (TypeError, ValueError) as exc:
			raise ValueError("min_score 必须是数字") from exc
		if not 0 <= minimum_score <= 100:
			raise ValueError("min_score 超出允许范围")
		conditions.append("score >= ?")
		params.append(minimum_score)
	created_within = str(filters.get("created_within") or "").strip()
	if created_within == "today":
		conditions.append("created_at >= datetime('now', 'localtime', 'start of day', 'utc')")
	elif created_within in {"3d", "7d"}:
		days = 3 if created_within == "3d" else 7
		conditions.append(f"created_at >= datetime('now', '-{days} days')")
	elif created_within:
		raise ValueError("created_within 参数无效")
	query = "SELECT * FROM jobs"
	if conditions:
		query += " WHERE " + " AND ".join(conditions)
	query += " ORDER BY created_at DESC, score DESC"
	rows = [dict(row) for row in conn.execute(query, params).fetchall()]

	salary_min = filters.get("salary_min")
	salary_max = filters.get("salary_max")
	try:
		minimum_salary = None if salary_min in (None, "") else float(salary_min)
		maximum_salary = None if salary_max in (None, "") else float(salary_max)
	except (TypeError, ValueError) as exc:
		raise ValueError("薪资范围必须是数字") from exc
	if minimum_salary is not None and maximum_salary is not None and minimum_salary > maximum_salary:
		raise ValueError("最低薪资不能高于最高薪资")
	if minimum_salary is None and maximum_salary is None:
		return rows
	filtered: list[dict[str, Any]] = []
	for row in rows:
		salary_range = parse_monthly_salary_k(str(row.get("salary") or ""))
		if salary_range is None:
			continue
		job_minimum, job_maximum = salary_range
		if minimum_salary is not None and job_maximum < minimum_salary:
			continue
		if maximum_salary is not None and job_minimum > maximum_salary:
			continue
		filtered.append(row)
	return filtered


def _selected_rows(
	conn: sqlite3.Connection,
	scope: str,
	job_ids: Iterable[Any] | None,
	filters: dict[str, Any] | None,
) -> list[dict[str, Any]]:
	scope = str(scope or "all").strip().lower()
	if scope not in {"all", "filtered", "selected"}:
		raise ValueError("导出范围无效")
	ids = _normalize_export_ids(job_ids)
	if scope == "all":
		return _filtered_rows(conn)
	if scope == "filtered":
		return _filtered_rows(conn, filters)
	if not ids:
		raise ValueError("所选岗位不能为空")
	placeholders = ",".join("?" for _ in ids)
	rows = [dict(row) for row in conn.execute(
		f"SELECT * FROM jobs WHERE deleted_at IS NULL AND id IN ({placeholders}) ORDER BY created_at DESC, score DESC",
		ids,
	).fetchall()]
	found = {str(row["id"]) for row in rows}
	invalid_ids = [job_id for job_id in ids if job_id not in found]
	if invalid_ids:
		raise InvalidJobSelectionError(invalid_ids)
	return rows


def _failure_reason(job: dict[str, Any]) -> str:
	value = job.get("score_failure_json")
	if value:
		try:
			payload = json.loads(str(value))
			if isinstance(payload, dict) and payload.get("message"):
				return str(payload["message"])[:240]
		except (TypeError, json.JSONDecodeError):
			pass
	reason = str(job.get("score_reason") or "")
	if reason.startswith(("AI评分失败:", "AI 评分失败:", "评分失败:")):
		return reason.split(":", 1)[1].strip()[:240]
	return ""


def _greeting_failure_reason(job: dict[str, Any]) -> str:
	value = job.get("greeting_failure_json")
	if value:
		try:
			payload = json.loads(str(value))
			if isinstance(payload, dict):
				return str(payload.get("user_message") or payload.get("message") or "")[:240]
		except (TypeError, json.JSONDecodeError):
			pass
	return ""


def _row_values(job: dict[str, Any]) -> list[Any]:
	platform = str(job.get("source_platform") or "boss").strip() or "boss"
	platform_labels = {"boss": "BOSS 直聘", "zhilian": "智联招聘", "51job": "前程无忧"}
	city_code = str(job.get("source_city_code") or "").strip()
	if not city_code and platform == "boss":
		city_code = str(job.get("city_code") or "").strip() or (get_city_code(str(job.get("city") or "")) or "")
	job = {
		**job,
		"source_platform_label": platform_labels.get(platform, platform),
		"platform_city_code": city_code,
		"recruitment_type_label": {"campus": "校招", "experienced": "社招"}.get(
			str(job.get("recruitment_type") or ""), "未识别"
		),
		"score_failure_reason": _failure_reason(job),
		"greeting_failure_reason": _greeting_failure_reason(job),
	}
	return [job.get(key, "") if job.get(key) is not None else "" for _, key in EXPORT_COLUMNS]


def _csv_safe(value: Any) -> str:
	text = "" if value is None else str(value)
	if text.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
		return "'" + text
	return text


def build_csv(rows: list[dict[str, Any]]) -> bytes:
	buffer = io.StringIO(newline="")
	writer = csv.writer(buffer, lineterminator="\r\n")
	writer.writerow([label for label, _ in EXPORT_COLUMNS])
	for job in rows:
		writer.writerow([_csv_safe(value) for value in _row_values(job)])
	return buffer.getvalue().encode("utf-8-sig")


def build_xlsx(rows: list[dict[str, Any]]) -> bytes:
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font
	except ImportError as exc:
		raise RuntimeError("导出 XLSX 需要安装 openpyxl") from exc
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "岗位池"
	headers = [label for label, _ in EXPORT_COLUMNS]
	sheet.append(headers)
	for cell in sheet[1]:
		cell.font = Font(bold=True)
	for job in rows:
		values = _row_values(job)
		sheet.append(values)
		url_column = next(index for index, (_, key) in enumerate(EXPORT_COLUMNS, start=1) if key == "url")
		cell = sheet.cell(sheet.max_row, url_column)
		if cell.value:
			cell.hyperlink = str(cell.value)
			cell.style = "Hyperlink"
	for index, (_, key) in enumerate(EXPORT_COLUMNS, start=1):
		max_length = max(len(str(sheet.cell(row, index).value or "")) for row in range(1, min(sheet.max_row, 30) + 1))
		sheet.column_dimensions[chr(64 + index) if index <= 26 else sheet.cell(1, index).column_letter].width = min(max(max_length + 2, 12), 48)
		if key in {"jd", "score_reason", "greeting"}:
			for row in range(2, sheet.max_row + 1):
				cell = sheet.cell(row, index)
				alignment = copy(cell.alignment)
				alignment.wrap_text = True
				cell.alignment = alignment
	sheet.freeze_panes = "A2"
	sheet.auto_filter.ref = sheet.dimensions
	output = io.BytesIO()
	workbook.save(output)
	return output.getvalue()


def export_jobs(
	conn: sqlite3.Connection,
	*,
	format: str = "xlsx",
	scope: str = "all",
	job_ids: Iterable[Any] | None = None,
	filters: dict[str, Any] | None = None,
) -> tuple[bytes, str, str]:
	format = str(format or "xlsx").strip().lower()
	if format not in {"csv", "xlsx"}:
		raise ValueError("导出格式只支持 xlsx 或 csv")
	rows = _selected_rows(conn, scope, job_ids, filters)
	if format == "csv":
		return build_csv(rows), "text/csv; charset=utf-8", "jobagent-jobs.csv"
	if format == "xlsx":
		return build_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "jobagent-jobs.xlsx"


def export_row_count(
	conn: sqlite3.Connection,
	*,
	scope: str = "all",
	job_ids: Iterable[Any] | None = None,
	filters: dict[str, Any] | None = None,
) -> int:
	return len(_selected_rows(conn, scope, job_ids, filters))
